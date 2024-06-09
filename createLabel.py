from fpdf import FPDF
import qrcode
import openpyxl
from pathlib import Path
import os
import time

import pandas as pd



def generate_qr_code(data, output_filename="my_qr_code.png"):

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=2,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)

    # Create QR code image
    qr_image = qr.make_image(fill_color="black", back_color="white")

    # Save the image
    qr_image.save(output_filename)


def delete_png_files_based_on_excel_column(excel_path):

    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Assuming the column index is fixed (e.g., column B)
        column_index = 1  # Adjust as needed

        # Get the folder path where the Excel file is located
        folder_path = os.path.dirname(excel_path)

        # Iterate through the values in the specified column
        for cell in sheet.iter_rows(min_row=2, min_col=column_index, values_only=True):
            value = cell[0]
            if value:
                # Construct the PNG file path
                png_file_path = os.path.join(folder_path, f"{value}.png")

                # Check if the PNG file exists and delete it
                if os.path.exists(png_file_path):
                    os.remove(png_file_path)

    except Exception as e:
        print(f"An error occurred: {str(e)}")

        

def read_excel_data(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]  # Replace with your sheet name
    data = []

    for row in sheet.iter_rows(min_row=2):  # Skip header row
        eid, oldSerial, newModel, ruta, cc = row[0], row[2], row[3], row[1], row[5]  # Adjust column indices #comment->27 #ruta -> 30
        data.append((eid.value, newModel.value, oldSerial.value, ruta.value, cc.value))

    return data

def file_exists_using_pathlib(file_path: str) -> bool:
    my_file = Path(file_path)
    return my_file.is_file()
def create_labels(data, excelName):
    # Create an FPDF object
    pdf = FPDF()
    pdf.add_page()

    # Set font (Arial, bold, size 16)
    cantUsers = len(data)
    loopsQuant = (cantUsers// 8) + 1
    it = 0
    i = 0
    while it <= loopsQuant:


        # Define label dimensions (adjust as needed)
        label_width = 80
        label_height = 45
        label_spacing = 15 # Space between labels
        labels_per_page = 8
        label_count = 0
        
        for row in range(4):
            for col in range(2):
                if i < cantUsers:
                # Extract data for the label (EID, Personnel, Comment, Ruta)
                    eid, newModel, oldSerial, ruta, cc = data[i]  # Adjust based on your data structure

                    # Generate QR code (use your existing function)
                    generate_qr_code(eid, output_filename=str(eid) + ".png")

                    x = col * (label_width + label_spacing)
                    y = row * (label_height + label_spacing)

                    #Create label content (adjust positions as needed)
                    pdf.set_font('Arial', size=10)
                    pdf.set_xy(x+8, y+20)
                    pdf.multi_cell(label_width + 5, label_height - 5, txt="__________________________________________", border=1)

                    # Add QR code image (adjust position)
                    pdf.image(str(eid) + ".png", x=x + 75, y=y + 22, w=15, h=15)
                    

                    pdf.set_font('Arial', "B", 11) #Old Serial
                    pdf.set_xy(x+9, y+23)
                    pdf.multi_cell(label_width+15, 17, txt=f"Old Serial: {oldSerial}", border=0)

                    pdf.set_font('Arial', "B", 15)
                    pdf.set_xy(x+9, y+18) #EID 
                    pdf.multi_cell(label_width+1, 12, txt=str(eid) , border=0)

                    pdf.set_font('Arial', size=11)
                    pdf.set_xy(x+9, y+26) #COMMENT / COST CENTER
                    pdf.multi_cell(label_width, 24, txt= f"CC: {cc}" , border=0)


                    pdf.set_font('Arial', size=11)
                    pdf.set_xy(x+9, y+35) #COMMENT / COST CENTER
                    pdf.multi_cell(label_width, 24, txt= newModel, border=0) 

                    pdf.set_font('Arial', size=24)
                    pdf.set_xy(x+76, y+43) #RO
                    pdf.multi_cell(label_width, 24, txt= "RO", border=0)                                       



                    pdf.set_font('Arial', 'B', 11)
                    pdf.set_xy(x+9, y+11) #Location
                    pdf.multi_cell(label_width + 10, 85, f"Location: {ruta}", border=0)

                    label_count += 1
                    if label_count >= labels_per_page:
                        pdf.add_page()  # Start a new page
                        label_count = 0
                    i += 1
        it += 1
    # Save the PDF
    pdf.output("Etiquetas " + str(excelName)+'.pdf', 'F')


def split_excel_by_column_value(input_excel_path, column_name):
    # Read the input Excel file
    df = pd.read_excel(input_excel_path)

    # Get unique values in the specified column
    unique_values = df[column_name].unique()

    # Get the directory where the Python script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Split the data based on unique values and save each subset to a separate Excel file
    for value in unique_values:
        subset_df = df[df[column_name] == value]
        output_file_path = os.path.join(script_dir, f"{value}.xlsx")
        subset_df.to_excel(output_file_path, index=False)
        



def separate_excel_files(root_file_path: str) -> None:


    script_directory = os.path.dirname(os.path.realpath(__file__))
    filePath = os.path.join(script_directory, root_file_path)

    if file_exists_using_pathlib(filePath):
        root_workbook = openpyxl.load_workbook(root_file_path)
        root_sheet = root_workbook.active

        # Get unique values from the specified column (e.g., column F)
        unique_values = set(root_sheet["B2:B" + str(root_sheet.max_row)])
        
        # Create separate Excel files for each unique value
        for value_cell in unique_values:
            value = value_cell[0].value
            print(value)
            
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            # Copy the header row from the root sheet
            header_row = root_sheet[1]
            for cell in header_row:
                new_sheet[cell.column_letter + "1"] = cell.value

            # Copy matching rows to the new sheet
            for row in root_sheet.iter_rows(min_row=2, values_only=True):
                
                if row[1] == value:  # Assuming column B contains the LOCATION values to match
                    
                    new_sheet.append(row)

            # Save the new workbook
            print(f"Saving to: {value}.xlsx")


        # Close the root workbook
        new_workbook.save(f"{value}.xlsx")
        root_workbook.close()
        print("Separate Excel files created and matching rows copied successfully!")
    else:
        print(f"The file {filePath} does not exist")


listOfPossibleRoutes = ['PP', 'MDQ', 'ROS', 'FED']

fileName = input("Enter the excel name: \n")

root_excel_file = fileName + ".xlsx"

# split_excel_by_column_value(root_excel_file, "Facility Description")



for route in listOfPossibleRoutes:
    if file_exists_using_pathlib(file_path=str(route)+'.xlsx'):
        create_labels(read_excel_data(filename=str(route)+'.xlsx'), str(route))
    else:
        print(str(route) + " does not exists in this filer\n")

delete_png_files_based_on_excel_column(root_excel_file)